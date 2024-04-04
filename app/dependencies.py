from statistics import stdev, mean, covariance, variance
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.legends import Legend
from reportlab.lib.validators import Auto
from flask_wtf import FlaskForm
from wtforms import FileField, SubmitField

# spreadsheet read in and database initialization object
class fundGroup:

    def __init__(self, filename):
        #spreadsheet read in
        self.wb = load_workbook(filename=filename)
        self.ws = self.wb.worksheets[0]
        # dictionary to hold fund name: data created, list intitialized to hold all fund names
        self.indexes = {}
        self.fundNames = []

        # set row as first row with relevant data
        row = 3
        # nested while loop initialized to fill fundnames and indexes with data from the spreadsheet
        # while loop terminates when the first empty cell is reached in the fund names column of the spreadsheet
        while self.ws["A" + str(row)].value != None :
            tempKey = self.ws["A" + str(row)].value
            # checks if the fund has been previosuly added
            if (tempKey not in self.fundNames):
                self.fundNames.append(tempKey)
                #intitializes a list that will hold tuples that represent each cell; fund name : [(return in %, compound return, datetime(date of return)), ...]
                self.indexes[tempKey] = []
                col = 2
                while self.ws[get_column_letter(col) + str(row)].value != None and (type(self.ws[get_column_letter(col) + str(row)].value) == int 
                or type(self.ws[get_column_letter(col) + str(row)].value) == float) :
                    self.indexes[tempKey].append(((self.ws[get_column_letter(col) + str(row)].value), (1+(self.ws[get_column_letter(col) + str(row)].value)/100), 
                        (self.ws[get_column_letter(col)+"2"].value)))
                    col +=1
            row += 1

        self.length = len(self.fundNames)

# calculates the avg, min, max standard deviations of all possible 1,3,5 year return windows of a given fund
class standardDevGraph():

    def __init__(self, returns):
        # uses the period creator object to create return windows using given return stream
        self.periods = periodCreator(returns)
        # uses the built in calc function to create the final list
        self.data = [self.calc(self.periods.oneYr), self.calc(self.periods.threeYr), self.calc(self.periods.fiveYr)]

    # calculates standard deviation of each return window and returns a tuple: (mean, min, max) of the calculated standard deviations
    def calc(self, customPeriod):
        stdDevs = []
        
        for lst in customPeriod:
            stdDevs.append(stdev(lst))

        return([round(mean(stdDevs),2), round(min(stdDevs),2), round(max(stdDevs),2)])
    
# compares the growth of an intitial $10K investment of two funds 
class growth10KGraph():

    def __init__(self, fund1, fund2):
        # maximum comparison length is set
        self.length = min(len(fund1), len(fund2))
        # compound return is isolated
        self.returns = [self.createReturns(fund1), self.createReturns(fund2)]
        # data is calculated from isolated returns
        self.data = [self.calc(self.returns[0]), self.calc(self.returns[1])]
        # the proper period of comparison is turned into a list readable by java and reversed to proper order
        self.period = [str(fund1[i][2])[:10] for i in range(self.length)][::-1]
    
    # iterates through fund and grabs comp. return, then reverses the list to proper order
    def createReturns(self, fund):
        temp = [fund[i][1] for i in range(self.length)]
        return(temp[::-1])

    # calculates change from each return
    def calc(self, returns):
        data = [10000]
        for ret in returns:
            data.append(data[-1]*ret)
        return(data)

# calculates the rolling rate of return for a fund and index and compares the two, over and underperformance is tracked
class battingAverage():
    def __init__(self, fund, index):
        self.length = min(len(fund), len(index))
        self.returns = [self.createReturns(fund), self.createReturns(index)]
        self.fundPeriods = periodCreator(self.returns[0])
        self.indexPeriods = periodCreator(self.returns[1])
        # the rolling rate of return is calculated for both fund and index
        self.rawDataFund = [self.calc(self.fundPeriods.createPeriod(3)), self.calc(self.fundPeriods.oneYr), self.calc(self.fundPeriods.threeYr), self.calc(self.fundPeriods.fiveYr)]
        self.rawDataIndex = [self.calc(self.indexPeriods.createPeriod(3)), self.calc(self.indexPeriods.oneYr), 
            self.calc(self.indexPeriods.threeYr), self.calc(self.indexPeriods.fiveYr)]
        # the respective rolling periods are compared and # of over/underperformances are counted
        self.data = [self.compare(self.rawDataFund[0], self.rawDataIndex[0]), self.compare(self.rawDataFund[1], self.rawDataIndex[1]), 
            self.compare(self.rawDataFund[2], self.rawDataIndex[2]), self.compare(self.rawDataFund[3], self.rawDataIndex[3])]
    
    #isolates compound return
    def createReturns(self, fund):
        return([fund[i][1] for i in range(self.length)])

    # return is calculated based on windows, this number is different from an average
    def calc(self, periodReturns):
        returns = []
        # the return windows are iterated through
        for period in periodReturns:
            length = len(period)
            # periods less than a year are dealt with differently
            if length >= 12 :
                power = int(length/12)
                ret = 1
                for x in period:
                    ret = ret*x
                returns.append(((ret**(1/power))-1)*100)

            else:
                ret = 1
                for x in period:
                    ret = ret*x
                returns.append((ret-1)*100)
        return(returns)

    # takes two lists of returns and compares each data point to each other, returns [# of periods, overperf., underperf., overperf.%]
    def compare(self, data1, data2):
        total = len(data1)
        under = 0
        over = 0
        for i in range(total):
            if data1[i] > data2[i]:
                over = over+1
            if data1[i] < data2[i]:
                under = under+1
        return([total, under, over, (over/total)*100])

# maps standard deviation with avg rolling period rate of return
class riskReturnGraph():

    def __init__(self, fund, index):
        # uses previous structure to acces rolling rate of return
        helper = battingAverage(fund, index)
        self.fundAvgs = [mean(helper.rawDataFund[1]), mean(helper.rawDataFund[2]), mean(helper.rawDataFund[3])]
        self.indexAvgs = [mean(helper.rawDataIndex[1]), mean(helper.rawDataIndex[2]), mean(helper.rawDataIndex[3])]

        indexReturns = [index[i][0] for i in range(len(index))]
        # uses previous structure to caculate standard deviation
        self.standards = standardDevGraph(indexReturns)

# calculates the difference in returns of a fund and an index over 1,3,5 year rolling periods
class excessReturnsGraph():

    def __init__(self, fund, index):
        # uses previous structure to get rolling periods of compound returns
        extension = battingAverage(fund, index)
        # does a calculation from batting average graph obeject's data
        self.data = [self.masterCalc(extension.rawDataFund[0], extension.rawDataIndex[0]), self.masterCalc(extension.rawDataFund[1], extension.rawDataIndex[1]),
        self.masterCalc(extension.rawDataFund[2], extension.rawDataIndex[2]), self.masterCalc(extension.rawDataFund[3], extension.rawDataIndex[3])]
    
    # calculates difference in returns and averages them
    def masterCalc(self, frates, irates):
        temp = []
        for i in range(len(frates)):
            temp.append(frates[i]-irates[i])
        return(round(mean(temp),2))

# calculates the mean,min,max of the betas of 1,3,5 year rolling periods
class betaGraph():

    def __init__(self, fund, index):
        self.length = min(len(fund), len(index))
        fundP = periodCreator(self.createReturns(fund))
        indexP = periodCreator(self.createReturns(index))
        # calculation is performed to create data
        self.data = [self.calc(fundP.oneYr, indexP.oneYr), self.calc(fundP.threeYr, indexP.threeYr), self.calc(fundP.fiveYr, indexP.fiveYr)]

    # list is created that will hold all calculated betas, beta calculation in helper method
    def calc(self, periodF, periodI):
        lst = []
        for i in range(len(periodF)):
            lst.append(self.beta(periodF[i], periodI[i]))
        return([round(mean(lst),2), round(min(lst),2), round(max(lst),2)])

    # beta is calculated if possible, if not 0 is returned
    def beta(self, retF, retI):
        try:
            return((covariance(retF, retI))/(variance(retI)))
        except:
            return(0)

    # isolates return in %
    def createReturns(self, fund):
        return([fund[i][0] for i in range(self.length)])

# creates all 1,3,5 year rolling periods of the given return stream w/ functionality to build any round-month length periods
class periodCreator:
    def __init__(self, returns):
        self.base = returns
        self.length = len(returns)
        self.oneYr = self.createPeriod(12)
        self.threeYr = self.createPeriod(36)
        self.fiveYr = self.createPeriod(60)

    # the returns are iterated through and all possible periods of length interval are stored in a list
    def createPeriod(self, interval):
        list = []
        for i in range((self.length-(interval-1)+1)):
            list.append((self.base[i:i+interval]))
        
        if len(list) == 0:
            list = [[0,0],[0,0]]

        return(list)
    
# File Upload Form Class
class UploadFileForm(FlaskForm):
    file = FileField("File")
    submit = SubmitField("Upload File")

# Market Value Aggregator methods
def createDict(ws):
    tickers = []
    noTick = []
    fundDict = {}
    accCol = ""
    tickCol = ""
    mvCol = ""

    for i in range(1, ws.max_column+1):
        temp = get_column_letter(i)
        if ws[temp+"1"].value == "Account Name":
            accCol = temp
        elif ws[temp+"1"].value == "Ticker":
            tickCol = temp
        else:
            mvCol = temp

    for i in range(2, ws.max_row + 1):
        accName = ws[accCol + str(i)].value
        tick = ws[tickCol + str(i)].value
        if tick not in tickers and accName not in noTick:
            if tick == "":
                noTick.append(accName)
                fundDict[accName] = [accName, ws[mvCol + str(i)].value]
                if fundDict[accName][1] == None:
                    fundDict[accName][1] = 0
            else:
                tickers.append(tick)
                fundDict[tick] = [accName, ws[mvCol + str(i)].value]
                if fundDict[tick][1] == None:
                    fundDict[tick][1] = 0
        else:
            try:
                if tick == "":
                        fundDict[accName][1] = fundDict[accName][1] + ws[mvCol + str(i)].value
                else:
                    fundDict[tick][1] = fundDict[tick][1] + ws[mvCol + str(i)].value
            except:
                pass
    
    return (fundDict)

def createReport(dict):
    outputWB = Workbook()
    active = outputWB.active
    active.cell(row = 1, column = 1).value = "Name"
    active.cell(row = 1, column = 2).value = "Ticker"
    active.cell(row = 1, column = 3).value = "Total Market Value"
    n = 2
    for key in dict: 
        active.cell(row = n, column = 1).value = dict[key][0]
        active.cell(row = n, column = 2).value = key
        active.cell(row = n, column = 3).value = dict[key][1]
        n = n + 1

    outputWB.save("app/static/files/output.xlsx")

"""PDF Creation Functions"""
def add_legend(draw_obj, chart, data):
    legend = Legend()
    legend.alignment = 'right'
    legend.x = 10
    legend.y = 70
    legend.colorNamePairs = Auto(obj=chart)
    draw_obj.add(legend)

# calculates the return of a portfolio consisting of x% equity and y% bonds
def returnStreams(market, bond, x, y):
    masterList = []
    for i in range(len(market)-34):
        masterList.append([(market[i]*x)+(bond[i]*y) for i in range(i, i+35)])
    return(masterList)

#initializes a 35 year return window with a flat rate of return of 5%
returns = {str(i+1): [0.05, False] for i in range(35)}

# two master lists are created storing returns from 1928-2021 for the equity market and for us bonds
marketReturns = [43.81, -8.30, -25.12, -43.84, -8.64, 49.98, -1.19, 46.74, 31.94, -35.34, 29.28, -1.10, -10.67, -12.77, 19.17, 25.06, 19.03, 35.82, -8.43, 5.20, 
    5.70, 18.30, 30.81, 23.68, 18.15, -1.21, 52.56, 32.60, 7.44, -10.46, 43.72, 12.06, 0.34, 26.64, -8.81, 22.61, 16.42, 12.40, -9.97, 23.80, 10.81, -8.24, 3.56, 
    14.22, 18.76, -14.31, -25.90, 37.00, 23.83, -6.98, 6.51, 18.52, 31.74, -4.70, 20.42, 22.34, 6.15, 31.24, 18.49, 5.81, 16.54, 31.48, -3.06, 30.23, 7.49, 9.97, 
    1.33, 37.20, 22.68, 33.10, 28.34, 20.89, -9.03, -11.85, -21.97, 28.36, 10.74, 4.83, 15.61, 5.48, -36.55, 25.94, 14.82, 2.10, 15.89, 32.15, 13.52, 1.38, 11.77, 
    21.61, -4.23, 31.21, 18.02, 28.47]

bondReturns = [0.84, 4.20, 4.54, -2.56, 8.79, 1.86, 7.96, 4.47, 5.02, 1.38, 4.21, 4.41, 5.40, -2.02, 2.29, 2.49, 2.58, 3.80, 3.13, 0.92, 1.95, 4.66, 0.43, -0.30, 
    2.27, 4.14, 3.29, -1.34, -2.26, 6.80, -2.10, -2.65, 11.64, 2.06, 5.69, 1.68, 3.73, 0.72, 2.91, -1.58, 3.27, -5.01, 16.75, 9.79, 2.82, 3.66, 1.99, 3.61, 15.98, 
    1.29, -0.78, 0.67, -2.99, 8.20, 32.81, 3.20, 13.73, 25.71, 24.28, -4.96, 8.22, 17.69, 6.24, 15.00, 9.36, 14.21, -8.04, 23.48, 1.43, 9.94, 14.92, -8.25, 16.66, 
    5.57, 15.12, 0.38, 4.49, 2.87, 1.96, 10.21, 20.10, -11.12, 8.46, 16.04, 2.97, -9.10, 10.75, 1.28, 0.69, 2.80, -0.02, 9.64, 11.33, -4.42]

# master dictionary of returns for all relevent portfolio compisitions
returnCombinations = {'market': returnStreams(marketReturns, bondReturns, 1, 0), 'bonds': returnStreams(marketReturns, bondReturns, 0, 1), 
    '90M10B': returnStreams(marketReturns, bondReturns, 0.9, 0.1), '80M20B': returnStreams(marketReturns, bondReturns, 0.8, 0.2), 
    '70M30B': returnStreams(marketReturns, bondReturns, 0.7, 0.3), '60M40B': returnStreams(marketReturns, bondReturns, 0.6, 0.4), 
    '50M50B': returnStreams(marketReturns, bondReturns, 0.5, 0.5), '40M60B': returnStreams(marketReturns, bondReturns, 0.4, 0.6), 
    '30M70B': returnStreams(marketReturns, bondReturns, 0.3, 0.7), '20M80B': returnStreams(marketReturns, bondReturns, 0.2, 0.8), 
    '10M90B': returnStreams(marketReturns, bondReturns, 0.1, 0.9)}
